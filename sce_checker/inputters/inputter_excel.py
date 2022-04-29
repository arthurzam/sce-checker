import functools
from itertools import count
from pathlib import Path
from typing import Iterable, NamedTuple, Tuple, TypeVar, Union

from openpyxl import load_workbook

from .. import consts


class ErrorComment(NamedTuple):
    value: float
    comment: str

    def __str__(self) -> str:
        if self.value == 0:
            return self.comment
        return f'{self.comment}{consts.HTML_INDENT}<span dir="ltr">({-self.value:+.2f})</span>'


T = TypeVar('T')
class StudentErrorHolder(dict[str, Union[T, 'StudentErrorHolder[T]']]):
    def append(self, titles: Iterable[str], err: ErrorComment) -> None:
        *high, low = titles
        items = self
        for title in high:
            assert isinstance(items := self.setdefault(title, StudentErrorHolder()), StudentErrorHolder)
        assert isinstance(items := items.setdefault(low, []), list)
        items.append(err)

    @staticmethod
    def __title_output(title_prefix: str, title: str) -> str:
        if not title:
            return ''
        elif not title_prefix:
            return f'{title}: '
        return f'{title_prefix} {title}: '

    def tostring(self, titles_prefixes: Tuple[str, ...], indent_level=1) -> str:
        if not self:
            return ''
        indent = f'{consts.HTML_NEWLINE}{indent_level * consts.HTML_INDENT}'
        if isinstance(next(iter(self.values())), StudentErrorHolder):
            return consts.HTML_NEWLINE.join(
                self.__title_output(titles_prefixes[0], title) + (
                    indent if len(parts) > 1 else ''
                ) + parts.tostring(titles_prefixes[1:], indent_level + int(len(parts) > 1))
                for title, parts in self.items()
            )
        else:
            return f'{consts.HTML_NEWLINE}{(indent_level - 1) * consts.HTML_INDENT}'.join(
                self.__title_output(titles_prefixes[0], title) + (
                    indent if len(parts) > 1 else ''
                ) + indent.join(map(str, parts))
                for title, parts in self.items()
            )


class InputterExcel:
    def __init__(self, excel_file: Path) -> None:
        self.sheet = load_workbook(str(excel_file), data_only=True).worksheets[0]

    def cell(self, row: int, col: int) -> str:
        value = self.sheet.cell(row=row, column=col).value
        return str(value) if value is not None else ''

    @functools.cached_property
    def row_total_marks(self) -> int:
        for row in count(consts.XLSM_ROW_NAMES):
            if self.cell(row, 1) == consts.XLSM_ROW_TITLE_TOTAL_MARKS:
                return row
        return -1

    @functools.cached_property
    def clear_char(self) -> str:
        for row in count(consts.XLSM_ROW_NAMES):
            if (value := self.cell(row, 1)) == consts.XLSM_ROW_TITLE_CLEAR_CHAR:
                return value
        return ''

    def students_cols(self) -> Iterable[int]:
        col = consts.XLSM_COL_NAMES_START
        while self.cell(consts.XLSM_ROW_NAMES, col) != '':
            yield col
            col += 1

    def __iter__(self) -> Iterable[consts.StudentGrade]:
        sections_range = range(1, consts.XLSM_COL_NAMES_START - 2)
        titles_prefixes = tuple(self.cell(consts.XLSM_ROW_NAMES, col) for col in sections_range)
        for col in self.students_cols():
            comment = StudentErrorHolder()
            for row in range(consts.XLSM_ROW_NAMES + 1, self.row_total_marks):
                if self.cell(row, col) in ('', self.clear_char):
                    continue
                # TODO: implement formatting
                try:
                    value = float(self.cell(row, consts.XLSM_COL_NAMES_START - 2))
                except ValueError:
                    pass # TODO: raise custom error
                titles = (self.cell(row, col) for col in sections_range)
                error = self.cell(row, consts.XLSM_COL_NAMES_START - 1)
                comment.append(titles, ErrorComment(value, error))
            yield consts.StudentGrade(
                name=self.cell(consts.XLSM_ROW_NAMES, col),
                grade=self.cell(self.row_total_marks, col),
                comment=comment.tostring(titles_prefixes),
            )
