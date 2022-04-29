import os
import shutil
from pathlib import Path
from typing import Iterable
from zipfile import ZipFile

from openpyxl import load_workbook

from . import consts


def _extract_files(files_zip: Path, files_dir: Path):
    shutil.rmtree(files_dir, ignore_errors=True)
    os.makedirs(files_dir)

    with ZipFile(files_zip) as zip_f:
        zip_f.extractall(files_dir)

def _rename_files(files_dir: Path) -> Iterable[str]:
    for name in files_dir.iterdir():
        student_name = consts.str_cleanup(name.name.split('_', 1)[0].strip())
        name.rename(files_dir / student_name)
        yield student_name

def _prepare_excel(dst_dir: Path, names: Iterable[str]):
    template = Path(__file__).parent / 'template.xlsm'
    with template.open('rb') as source:
        workbook = load_workbook(source, read_only=False, keep_vba=True, data_only=False)
        worksheet = workbook.worksheets[0]
        for column, name in enumerate(names, start=consts.XLSM_COL_NAMES_START):
            worksheet.cell(row=consts.XLSM_ROW_NAMES, column=column).value = name
        workbook.save(str(dst_dir / 'check.xlsm'))

def prepare_check_xlsm(files_zip: Path, dst_dir: Path) -> int:
    files_dir = dst_dir / 'files'
    _extract_files(files_zip, files_dir)
    _prepare_excel(dst_dir, names := sorted(_rename_files(files_dir)))
    return len(names)
